from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import openpyxl
import os
from collections import Counter

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

MENU = {
    "Pizza": 250,
    "Burger": 150,
    "Pasta": 200,
    "Tacos": 180,
    "Sandwich": 140,
    "Fries":100
}

EXCEL_FILE = "orders.xlsx"

# Create Excel if not exists
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Customer Name", "Food Item", "Quantity", "Total Price"])
    wb.save(EXCEL_FILE)


def get_trending():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        items.append(row[1])

    counter = Counter(items)
    top3 = counter.most_common(3)
    return [item[0] for item in top3]


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    trending = get_trending()
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "menu": MENU, "trending": trending}
    )


@app.post("/order")
async def place_order(
    name: str = Form(...),
    food: str = Form(...),
    quantity: int = Form(...)
):
    if food not in MENU:
        return JSONResponse(
            content={"error": "Please enter from menu only."},
            status_code=400
        )

    price = MENU[food]
    total = price * quantity

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    sheet.append([name, food, quantity, total])
    wb.save(EXCEL_FILE)

    return {"total": total}

@app.get("/top-trending")
def top_trending():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        items.append(row[1])

    if not items:
        return {"top": None}

    counter = Counter(items)
    top_item = counter.most_common(1)[0][0]

    return {"top": top_item}
